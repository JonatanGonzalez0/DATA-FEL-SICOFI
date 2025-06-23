from Item import Item, Compra
from operacion import Operacion
from datetime import datetime
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *
from tkinter import filedialog,messagebox,ttk
import xml.etree.ElementTree as ET
import webbrowser
import pandas as pd
import re,os
from os.path import expanduser
import xlrd
import openpyxl
import math


#for style update pip install --upgrade xlsxwriter

home = expanduser("~")
Documents_location = os.path.join(home,'Documents')

path_DataOutput = Documents_location + "\\FEL-A-SICOFI"

path_DataOutputCompras = Documents_location + "\\FEL-A-SICOFI\\COMPRAS"
path_DataOutputVentas = Documents_location + "\\FEL-A-SICOFI\\VENTAS"

boolProcesar = False
nombreReceptor = ""
procesarSiguiente = True


if not os.path.exists(path_DataOutput):
    os.makedirs(path_DataOutput)
    print(f'No existia el directorio  {path_DataOutput} se ha creado')


if not os.path.exists(path_DataOutputCompras):
    os.makedirs(path_DataOutputCompras)
    print(f'No existia el directorio  {path_DataOutputCompras} se ha creado')

if not os.path.exists(path_DataOutputVentas):
    os.makedirs(path_DataOutputVentas)
    print(f'No existia el directorio  {path_DataOutputVentas} se ha creado')

nombreEmisor = ""
nomenclatura = ('52000000-GASTOS SOBRE COMPRAS',
                                '52010000-Gastos De Almacenaje',
                                '52020000-Honorarios Aduanales',
                                '52030000-Gastos De Importacion',
                                '52040000-Peaje',
                                '52050000-Fletes  Y  Acarreos',
                                '52060000-Otros Gastos De Compras',
                                '52070000-Devoluciones Y Rebajas / Compras',
                                '52080000-Inventario Final',
                                '52090000-Repuestos Y Accesorios',
                                '52100000-Gastos De Mantenimiento',
                                '52110000-Cuotas Varias',
                                '52120000-Gastos De Vehículo',
                                '52130000-Combustibles Y Lubricantes',
                                '52140000-Viaticos',
                                '52150000-Publicidad',
                                '52160000-Gastos  De Viaje',
                                '52170000-Compra De Gas Propano',
                                '52180000-Oxigeno Para Soldar',
                                '52190000-Fumigación',
                                '52200000-Marchamo',
                                '52210000-Alquiler De Maquinaria',
                                '52220000-Energia Electrica',
                                '52230000-Servicios Contratados',
                                '53000000-GASTOS DE VENTAS',
                                '53010000-Comisiones Sobre Ventas',
                                '53020000-Repuestos Y Accesorios',
                                '53030000-Gastos De Mantenimiento',
                                '53040000-Cuotas Varias',
                                '53050000-Gastos De Vehiculos',
                                '53060000-Combustibles Y Lubricantes',
                                '53070000-Energia Electrica',
                                '53080000-Gastos De Representacion',
                                '53090000-Viaticos',
                                '53100000-Fletes  Y  Acarreos',
                                '53110000-Publicidad',
                                '53120000-Gastos De Viaje',
                                '53130000-Otros',
                                '53140000-Servicios Contratados',
                                '53150000-Compra De Ropa Extraviada',
                                '53160000-Tele-escucha',
                                '53170000-Compra De Gas Propano',
                                '53180000-Gastos De Tienda',
                                '53190000-Oxigeno Para Soldar',
                                '53200000-Sueldos Y Salarios  Sala Ventas',
                                '53210000-Combustibles Y Lubricantes Usa',
                                '53220000-Fertilizante',
                                '54000000-GASTOS DE ADMINISTRACION',
                                '54010000-Sueldos Y Salarios',
                                '54020000-Bonificacion Especial',
                                '54030000-Aguinaldo',
                                '54040000-Vacaciones',
                                '54050000-Indemnizacion',
                                '54060000-Beneficio A Empleados',
                                '54070000-Transporte A Empleados',
                                '54080000-Cuota Patronal',
                                '54090000-Impuestos Y Contribuciones',
                                '54100000-Seguros',
                                '54110000-Honorarios Profecionales',
                                '54120000-Honorarios Contables',
                                '54130000-Telefonos',
                                '54140000-Fotocopias',
                                '54150000-Utiles Y Enseres De Limpieza',
                                '54160000-Combustibles Y Lubricantes',
                                '54170000-Extraccion De Basura',
                                '54180000-Mantenimiento A Equipo',
                                '54190000-Seguridad',
                                '54200000-Alquileres',
                                '54210000-Papeleria Y Utiles',
                                '54220000-Suscripcion',
                                '54230000-Depreciacion',
                                '54240000-Amortizacion',
                                '54250000-Depreciacion Vehiculos',
                                '54260000-Depreciacion Mobiliario Y Equipo',
                                '54270000-Depreciacion Herramientas Y Equipo',
                                '54280000-Gastos Generales',
                                '54290000-Depreciacion Equipo De Computacion',
                                '54300000-Publicaciones',
                                '54310000-Isr Mensual',
                                '54320000-VIATICOS',
                                '54320100-Viaticos Para Combustibles',
                                '54330000-Repuestos Y Accesorios',
                                '54340000-Bono 14',
                                '54350000-Gastos Medicos',
                                '54360000-Parqueo',
                                '54370000-Dietas',
                                '54380000-Atencion A Clientes',
                                '54390000-Agua',
                                '54400000-Depreciacion Construccion',
                                '54410000-Peaje',
                                '54420000-Cuentas Incobrables',
                                '54430000-Donaciones',
                                '54440000-Ofrendas A Ministerios',
                                '54450000-Mantenimiento De Casa Pastoral',
                                '54460000-Becas De Estudio',
                                '54470000-Diezmo De Diezmo',
                                '54480000-Ofrenda De Amor',
                                '54490000-Traslado De Fondos Cta. Fiscal',
                                '54500000-Refacciones Por Cursos Impartidos',
                                '54510000-Gastos Fotos Por Cursos',
                                '54520000-Eventos Evangelicos',
                                '54530000-Gastos De Oficina',
                                '54540000-Gastos Areas De Niños',
                                '54550000-Gastos Area Diaconado',
                                '54560000-Programa De Radio',
                                '54570000-Material Didactico',
                                '54580000-Gastos De Area De Alabanza',
                                '54590000-Gastos De Librería',
                                '54600000-Fabricacion De Uniformes',
                                '54610000-Sala Cuna',
                                '54620000-Matrimonios',
                                '54630000-Encuentro Niños',
                                '54640000-Encuentro Adultos',
                                '54650000-Retiro Pastores',
                                '54660000-Gastos De Construccion',
                                '54670000-Retiros',
                                '54680000-Retiro Señoritas',
                                '54690000-Retiro Jovenes',
                                '54700000-Retiro Hombres',
                                '54710000-Retiro Mujeres',
                                '54720000-Clasificados',
                                '54730000-Santa Cena',
                                '54740000-Diaconado',
                                '54750000-Encuentro Señoritas',
                                '54760000-Encuentro Señoras',
                                '54770000-Area De Computo',
                                '54780000-Traslado Cuenta De Ahorro',
                                '54790000-Reparacion Y Mantenimiento',
                                '54800000-Correo Y Encomiendas',
                                '54810000-Impresión De Facturas',
                                '54820000-Medicamentos',
                                '54830000-Celulares Pastores',
                                '54840000-Refacciones',
                                '54850000-Pasajes',
                                '54870000-VIGILIAS',
                                '54870100-Vigilia Area De Jovenes',
                                '54870200-Vigilia Mujeres',
                                '54870300-Vigilia Hombres',
                                '54870400-Retenciones De Isr Por Intereses',
                                '60000000-OTROS GASTOS Y PRODUCTOS FINANCIEROS',
                                '61000000-GASTOS FINANCIEROS',
                                '61010000-Intereses Bancarios',
                                '61020000-Descuentos Concedidos',
                                '61030000-Comisiones Bancarias',
                                '61040000-Intereses Otros',
                                '61050000-Intereses Pagados',
                                '62000000-GASTOS NO DEDUCIBLES',
                                '62010000-Gastos No Deducibles')

app = tk.Tk()
type_service_bien = ttk.Combobox(app,font=("Courier", 20))

def carga_Menu():
    Carpeta_Raiz = os.path.dirname(os.path.abspath(__file__))
    app.configure(bg='#003153')
    ancho_ventana = 600
    alto_ventana = 600
    #Centrar app dependiendo del monitor
    x_ventana = app.winfo_screenwidth() // 2 - ancho_ventana // 2
    y_ventana = app.winfo_screenheight() // 2 - alto_ventana // 2
    app.title("Data Analizer FEL-Sicofi Made by Jonatan")
    posicion = str(ancho_ventana) + "x" + str(alto_ventana) + "+" + str(x_ventana) + "+" + str(y_ventana)
    app.geometry(posicion)

    menubar = tk.Menu(app)
    filemenu = tk.Menu(menubar,tearoff=0)

    etiquetaM = tk.Label(app,text="BIENVENIDO",font=("Courier", 45))
    etiquetaM.pack(padx=30,pady=5)
    etiquetaM.configure(bg='#003153',fg='#DADEE7')
    filemenu.add_command(label="Salir", command=salir)

    menubar.add_cascade(label="Menu", menu=filemenu)
    app.config(menu=menubar)

    cargarBotones()
    app.mainloop()

def cargarBotones(): 
    global type_service_bien
    type_service_bien['values']=('Servicios afectos','Bienes afectos')
    type_service_bien.current(0)
    type_service_bien.pack(pady=20)
    
    combostyle = ttk.Style()
    type_service_bien['state'] = 'readonly'
    combostyle.theme_create('combostyle', parent='alt',
                         settings = {'TCombobox':
                                     {'configure':
                                      {'selectbackground': '#66b2b2',
                                       'fieldbackground': '#66b2b2',
                                       'background': '#66b2b2'
                                       }}}
                         )
    combostyle.theme_use('combostyle') 

    btn_cargarArchivo = tk.Button(app,text="VENTAS EXCEL A SICOFI",font=("Courier", 20),command=cargarArchivo)
    btn_cargarArchivo.configure(background='#66b2b2',fg='#fdfbfb')
    btn_cargarArchivo.pack( pady=20)

    # SEPARADOR DE BOTONES
    separador = ttk.Separator(app, orient='horizontal')
    separador.pack(fill='x', padx=20, pady=10)


    btn_buscarCarpeta = tk.Button(app,text="ABRIR CARPETA ARCHIVOS",font=("Courier", 20),command=buscarCarpeta)
    btn_buscarCarpeta.configure(background='#66b2b2',fg='#fdfbfb')
    btn_buscarCarpeta.pack( pady=30)

    # SEPARADOR DE BOTONES
    separador = ttk.Separator(app, orient='horizontal')
    separador.pack(fill='x', padx=20, pady=10)
    
    btn_ExtraerInfo = tk.Button(app,text="XML VENTAS a INVENTARIO",font=("Courier", 20),command=extraerInfoVentas)
    btn_ExtraerInfo.configure(background='#ff8c00',fg='#fdfbfb')
    btn_ExtraerInfo.pack( pady=20)

    btn_ExtraerInfo = tk.Button(app,text="XML COMPRAS a SICOFI",font=("Courier", 20),command=extraerInfoCompras)
    btn_ExtraerInfo.configure(background='#8823DB',fg='#fdfbfb')
    btn_ExtraerInfo.pack( pady=20)
   
def buscarCarpeta():
    os.startfile(path_DataOutput)
    
def salir():
    app.destroy()

def cargarArchivo():
    global type_service_bien
    try:
        Carpeta_Raiz = os.path.dirname(os.path.abspath(__file__))
        file_names = filedialog.askopenfilenames(initialdir=Carpeta_Raiz,title = "Selecciona uno o varios archivos")
        list_file_names = list(file_names)
        
        for file_name in list_file_names:
            
            #si archivo es xls convert to xlsx
            
            file_name = prepareFileExcel(file_name)
        
            df = pd.read_excel(file_name)
            serie = 'FEL'
            numero=0
            length_df = len(df.index)

            file_name_export = os.path.basename(file_name).split(".")[0]
            file = path_DataOutput + "\\"+ file_name_export + '.csv'

            with open(file,'w') as f:
                headers="serie,numero,ano,mes,dia,nit,nombre,servaf,servnaf,bienaf,biennaf,otros,inguat,status,tipodoc,concepto,descrip,nretenisr,vretenisr,tduafauca,nduafauca,tconstan,nconstan,vconstan,centro,seriefel,numerofel\n"
                f.write(headers)

                for index_df in range(length_df):
                    fechaDTE = df.loc[index_df]['Fecha de emisión']
                    matchFecha = re.search(r'\d{4}-\d{2}-\d{2}',fechaDTE).group()
                    strFecha = matchFecha.split("-")

                    ano = strFecha[0]
                    mes = strFecha[1]
                    dia = strFecha[2]

                    nit = str(df.loc[index_df]['ID del receptor'])
                    if nit =="CF":
                        nit ="C/F"
                    else:
                        size_nit = len(nit)
                        first_sub_string = nit[0:size_nit-1]
                        end_sub_string = nit[-1]
                        nit = first_sub_string + "-" + end_sub_string
                    
                    nombre = df.loc[index_df]['Nombre completo del receptor']
                    nombre = nombre.replace(",",'')

                    if type_service_bien.get()=="Servicios afectos":
                        try :
                            servaf = round(df.loc[index_df]['Monto (Gran Total)'],2)
                        except:
                            try:
                                servaf = round(df.loc[index_df]['Gran Total (Moneda Original)'],2)
                            except:
                                # crear excepcion si no se encuentra el campo 
                                servaf = 0

                        
                        servnaf = 0 

                        bienaf = 0
                        biennaf = 0
                    elif type_service_bien.get()=="Bienes afectos":
                        servaf = 0
                        servnaf = 0
                        try:
                            
                            bienaf = round(df.loc[index_df]['Monto (Gran Total)'],2)
                        except:
                            try:
                                bienaf = round(df.loc[index_df]['Gran Total (Moneda Original)'],2)
                            except:
                                # crear excepcion si no se encuentra el campo 
                                bienaf = 0
                        
                        biennaf = 0
                    
                    otros = 0
                    inguat = 0 
                    try:
                        estado = df.loc[index_df]['Marca de anulado ']
                    except:
                            estado = df.loc[index_df]['Marca de anulado']

                    if estado =="No":
                        status = 0
                    elif estado =='Si':
                        status = 1
                    
                    tipodoc = 2
                    concepto = 1

                    descrip = "Ventas varias"

                    nretenisr = ''
                    vretenisr = ''
                    tduafauca = ''
                    nduafauca = ''
                    tconstan = ''
                    nconstan = ''
                    vconstan = '' 
                    centro =  '' 

                    seriefel = df.loc[index_df]['Serie']

                    numerofel = df.loc[index_df]['Número del DTE']

                    strFila = '{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{}\n'.format(serie,numero,ano,mes,dia,nit,nombre,servaf,servnaf,bienaf,biennaf,otros,inguat,status,tipodoc,concepto,descrip,nretenisr,vretenisr,tduafauca,nduafauca,tconstan,nconstan,vconstan,centro,seriefel,numerofel)
                    f.write(strFila)
                
                f.close()
        messagebox.showinfo("Correcto","Se analizaron las facturas, se encuentraen en carpeta Documentos/FEL-A-SICOFI")
        webbrowser.open(path_DataOutput)    
    except Exception as e:
        print(e)
        messagebox.showwarning("Error","No se pudo realizar la operacion \n Error : [ " + str(e) + " ]")
        return

def extraerInfoCompras():
    global boolProcesar, nombreReceptor, nomenclatura, procesarSiguiente
    procesarSiguiente = True
    try:
        path_downloads = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads')     
        file_names = filedialog.askopenfilenames(initialdir=path_downloads,title = "Selecciona uno o varios archivos")
        list_file_names = list(file_names)
        compras =[]
        compras_no_procesadas = []
        nulls=[]
        download_errors = []
        ns = {'DTE':'http://www.sat.gob.gt/dte/fel/0.2.0'}
        for file_name in list_file_names:
            if not procesarSiguiente:
                break

            #si archivo solo contiene la palabra "null" no se procesa
            if os.path.getsize(file_name) == 4:
                #extract file name without extension
                file_name_export = os.path.basename(file_name).split(".")[0]
                nulls.append(file_name_export)
                continue
            try:
                try:
                    tree = ET.parse(file_name,parser=ET.XMLParser(encoding="utf-8"))
                except Exception as e:
                    print(e)
                    messagebox.showwarning(f"Error","El documento {file_name_export} se debe descargar de nuevo del portal de la SAT")
                    download_errors.append(file_name_export)
                    continue
                tree = ET.parse(file_name,parser=ET.XMLParser(encoding="utf-8"))
                ET.indent(tree, space="\t", level=0)
                tree.write(file_name, encoding="utf-8")     
                tree = ET.parse(file_name ,parser=ET.XMLParser(encoding="utf-8"))
                root = tree.getroot()  

                try:
                    DATOS_Emision = root.find('DTE:SAT',ns).find('DTE:DTE',ns).find('DTE:DatosEmision',ns)
                    DATOS_Certificacion = root.find('DTE:SAT',ns).find('DTE:DTE',ns).find('DTE:Certificacion',ns)
                except:
                    continue

                nombreEmisor = DATOS_Emision.find('DTE:Emisor',ns).get('NombreEmisor')
                nombreEmisor = nombreEmisor.replace(",","")
                nombreEmisor = nombreEmisor.replace("?","")

                nombreReceptor = DATOS_Emision.find('DTE:Receptor',ns).get('NombreReceptor')
                nombreReceptor = nombreReceptor.replace(",","")
                nombreReceptor = nombreReceptor.replace("?","")

                afiliacionIVA = DATOS_Emision.find('DTE:Emisor',ns).get('AfiliacionIVA')
                peqcont =""
                if afiliacionIVA == 'PEQ':
                    peqcont = "S"

                codigo_sucursal = DATOS_Emision.find('DTE:Emisor',ns).get('CodigoEstablecimiento')

                nitEmisor = DATOS_Emision.find('DTE:Emisor',ns).get('NITEmisor')
                # agregar guion al nit en el penultimo digito 1234569 -> 123456-9
                size_nit = len(nitEmisor)
                first_sub_string = nitEmisor[0:size_nit-1]
                end_sub_string = nitEmisor[-1]
                nitEmisor = first_sub_string + "-" + end_sub_string

                Datosgen = DATOS_Emision.find('DTE:DatosGenerales',ns)
                fecha = Datosgen.get('FechaHoraEmision')
                fecha = fecha.split("-")
                anio = fecha[0]
                mes = fecha[1]
                dia = fecha[2].split('T')[0]
                fecha = ('{}-{}-{}'.format(anio, mes, dia))  # Modified line
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()

                mesYear = fecha_obj.strftime('%m-%Y')

                tipoFact = Datosgen.get('Tipo')
                TipoDoc=3
                strTipoDoc = ''
                # TipoDoc: 1:Es factura normal, 2:FACE, 3:FEL, 4:Factura Especial, 5:Nota de débito, 6:Nota de crédito
                if tipoFact == 'FESP':
                    TipoDoc = 4
                    strTipoDoc = 'Factura Especial'
                elif tipoFact == 'RECI':
                    # SALTAR SEGUIR CON EL SIGUIENTE ARCHIVO
                    continue
                    
                elif tipoFact == 'NDEB':
                    TipoDoc = 5
                    strTipoDoc = 'Nota de débito'
                elif tipoFact == 'NCRE':
                    TipoDoc = 6
                    strTipoDoc = 'Nota de crédito'
                else:
                    strTipoDoc = 'Factura Electrónica'

                #Periodo: Periodo para contabilizar.  AAAAMM
                periodo = 0000
                periodo = anio + mes

                #certificacion
                noDTE = DATOS_Certificacion.find('DTE:NumeroAutorizacion',ns).get('Numero')
                serie = DATOS_Certificacion.find('DTE:NumeroAutorizacion',ns).get('Serie')

                #items
                items = DATOS_Emision.find('DTE:Items',ns)

                itemsTemp = []
                grandTotal = 0
                descripcionTemp = ''
                for item in items:
                    tipo = item.get('BienOServicio')
        
                    cant = item.find('DTE:Cantidad',ns).text
                    cant = int(float(cant))
                    descripcion = item.find('DTE:Descripcion',ns).text
                    descripcion = descripcion.replace('\n','')
                    descripcion = descripcion.replace('\t','')
                    descripcion = descripcion.replace(',','')
                    descripcionTemp +=" | " + descripcion + " | "

                    precioUnit = round(float(item.find('DTE:PrecioUnitario',ns).text),2)
                    total = round(float(item.find('DTE:Total',ns).text),2)

                    impuestos = item.find('DTE:Impuestos',ns)
                    impuesto = impuestos.find('DTE:Impuesto',ns)

                    monto_impuesto = round(float(impuesto.find('DTE:MontoImpuesto',ns).text),2)

                    bien_af = bien_naf = serv_af = serv_naf = 0
                    if tipo == 'B':
                        bien_af = total if monto_impuesto > 0 else 0
                        bien_naf = 0 if monto_impuesto > 0 else total
                    elif tipo == 'S':
                        serv_af = total if monto_impuesto > 0 else 0
                        serv_naf = 0 if monto_impuesto > 0 else total

                    itemTemp = Item(cant,descripcion,precioUnit,total,tipo,serv_af,serv_naf,bien_af,bien_naf,monto_impuesto)
                    itemsTemp.append(itemTemp)
                    grandTotal += total
                    
                
                # mostrar al usuario la informacion del documento en otra ventana y una tabla de listado de items y un boton si desea 
                # procesar la informacion o no

                #crear ventana
                ventana = tk.Toplevel(app)
                ventana.title("Informacion de Factura")
                ancho_ventana = 1024
                alto_ventana = 768
                #Centrar app dependiendo del monitor
                x_ventana = app.winfo_screenwidth() // 2 - ancho_ventana // 2
                y_ventana = app.winfo_screenheight() // 2 - alto_ventana // 2
                posicion = str(ancho_ventana) + "x" + str(alto_ventana) + "+" + str(x_ventana) + "+" + str(y_ventana)
                ventana.geometry(posicion)
                ventana.configure(bg='#003153')

                txtBoxM = tk.Text(ventana, font=("Coustard", 18))
                txtBoxM.insert("1.0", "Tipo de documento: "+ strTipoDoc)
                txtBoxM.insert("2.0", "\nFecha: " + fecha_obj.strftime('%d/%m/%Y'))
                txtBoxM.insert("3.0", "\nEmisor: " + nombreEmisor)
                txtBoxM.insert("4.0", "\nSerie: " + serie + " DTE: " + noDTE )
                txtBoxM.insert("5.0", "\nTotal Factura: Q" + str(round(grandTotal,2)))

                txtBoxM.configure(state='disabled', bg='#003153', fg='#DADEE7',width=70, height=5)

                # Coloca el widget de texto en las coordenadas deseadas
                txtBoxM.place(x=50, y=40)

                etiquetaItems = tk.Label(ventana, text="DETALLE COMPRA", font=("Courier", 20, 'bold'))
                etiquetaItems.place(x=370, y=200)
                etiquetaItems.configure(bg='#003153', fg='#DADEE7')

                # Crear un estilo
                stylettk = ttk.Style()
                stylettk.configure("Treeview", font=("Courier", 12))
                stylettk.configure("Treeview.Heading", font=("Courier", 14, 'bold'))
                #  Configurar los colores de fondo de las filas
                stylettk.configure('Treeview', rowheight=20)  # Ajustar la altura de las filas si es necesario
                stylettk.configure('evenrow.Treeview', background='white')  # Filas pares
                stylettk.configure('oddrow.Treeview', background='lightgrey')  # Filas impares

                # Crear el treeview
                tree = ttk.Treeview(ventana, columns=("Tipo","Cantidad", "Descripcion", "Precio Unitario", "Total", "Impuesto"), show="headings", style="Treeview",height=12)
                tree.place(x=30, y=240)

                # Definir los encabezados de las columnas
                tree.heading("Tipo", text="Tipo")
                tree.heading("Cantidad", text="Cantidad")
                tree.heading("Descripcion", text="Descripcion")
                tree.heading("Precio Unitario", text="Precio Unitario")
                tree.heading("Total", text="Total")
                tree.heading("Impuesto", text="Impuesto")

                # Ajustar el ancho de la columna "Descripcion"
                tree.column("Tipo", width=100, minwidth=100, anchor="center")
                tree.column("Cantidad", width=100, minwidth=100, anchor="center")
                tree.column("Descripcion", width=400, minwidth=250)
                tree.column("Precio Unitario", width=150, minwidth=100, anchor="center")
                tree.column("Total", width=100, minwidth=100, anchor="center")
                tree.column("Impuesto", width=100, minwidth=100, anchor="center")
                
                # Agregar los items al Treeview
                for i, item in enumerate(itemsTemp):
                    tipo = "Bien" if item.tipo == "B" else "Servicio"
                    # Usar colores de fondo alternos para las filas
                    if i % 2 == 0:
                        tree.insert("", "end", values=(tipo, item.cantidad, item.descripcion, "Q"+ str(round(item.precio,2)), "Q"+str(round(item.total,2)), "Q"+str(round(item.impuesto,2))), tags=('evenrow',))
                    else:
                        tree.insert("", "end", values=(tipo, item.cantidad, item.descripcion, "Q"+ str(round(item.precio,2)), "Q"+str(round(item.total,2)), "Q"+str(round(item.impuesto,2))), tags=('oddrow',))
                
                
                def procesarInformacion():
                    global boolProcesar
                    #verificar combo box para obtener el concepto con el current
                    concepto = c1.current() + 1
                    
                    #obtener cuenta seleccionada del combo box c2
                    cuenta = c2.get()
                    ctagasto = cuenta.split("-")[0]
                    serv_af = 0
                    comp_af = 0
                    serv_naf = 0
                    serv_afe = 0
                    serv_nafe = 0
                    comp_naf = 0
                    comp_afe = 0
                    comp_nafe = 0
                    otros = 0
                    
                    for item in itemsTemp:
                        if item.tipo == 'B':
                            bien_af = item.total if item.impuesto > 0 else 0
                            bien_naf = 0 if item.impuesto > 0 else item.total
                            comp_af += bien_af
                            comp_naf += bien_naf
                        elif item.tipo == 'S':
                            serv_af += item.total if item.impuesto > 0 else 0
                            serv_naf += 0 if item.impuesto > 0 else item.total                         
                        

                    if TipoDoc == 4:
                        messagebox.showwarning("Error","No se puede procesar facturas especiales\n Se creara un archivo de facturas no procesadas")
                        compras_no_procesadas.append(Compra(nitEmisor,nombreEmisor,peqcont,'',serie,noDTE,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descripcionTemp,TipoDoc,'',ctagasto,codigo_sucursal,'','','','','','',fecha_obj) )
                        ventana.destroy()
                        return
                    elif TipoDoc == 5:
                        messagebox.showwarning("Error","No se puede procesar notas de debito\n Se creara un archivo de facturas no procesadas")
                        compras_no_procesadas.append(Compra(nitEmisor,nombreEmisor,peqcont,'',serie,noDTE,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descripcionTemp,TipoDoc,'',ctagasto,codigo_sucursal,'','','','','','',fecha_obj) )
                        ventana.destroy()
                        return
                    elif TipoDoc == 6:
                        compras_no_procesadas.append(Compra(nitEmisor,nombreEmisor,peqcont,'',serie,noDTE,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descripcionTemp,TipoDoc,'',ctagasto,codigo_sucursal,'','','','','','',fecha_obj) )
                        messagebox.showwarning("Error","No se puede procesar notas de credito\n Se creara un archivo de facturas no procesadas")
                        ventana.destroy()
                        return
                    
                                 # compra nit,razon,peqcont,retenerisr,serie,numero,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descrip,tipodoc,centro,ctagasto,sucursal,isrreten,fchaisr,visrreten,isrtipo,isrconcep,isrbase
                    compras.append(Compra(nitEmisor,nombreEmisor,peqcont,'',serie,noDTE,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descripcionTemp,TipoDoc,'',ctagasto,codigo_sucursal,'','','','','','',fecha_obj))
                    ventana.destroy()
                
                def noProcesarInformacion():
                    global boolProcesar
                    ventana.destroy()
                    boolProcesar = False

                def terminarProcesamiento():
                    global procesarSiguiente
                    procesarSiguiente = False
                    ventana.destroy()

                btn_procesar = tk.Button(ventana, text="Procesar FACTURA", font=("Courier", 14), command=procesarInformacion)
                btn_procesar.configure(background='#80B628', fg='#fdfbfb')
                btn_procesar.place(x=250,y=680)

                btn_no_procesar = tk.Button(ventana, text="No Procesar", font=("Courier", 14), command=noProcesarInformacion)
                btn_no_procesar.configure(background='#B62828', fg='#fdfbfb')
                btn_no_procesar.place(x=450,y=680)

                btn_terminardeprocesar = tk.Button(ventana, text="Dejar de Procesar facturas", font=("Courier", 14), command=terminarProcesamiento)
                btn_terminardeprocesar.configure(background='#B62828', fg='#fdfbfb')
                btn_terminardeprocesar.place(x=650,y=680)


                #busqueda de la cuenta
                search_var = tk.StringVar()

                # Create a function to update the combobox
                def update_combobox(*args):
                    search_term = search_var.get()
                    # Find the closest match in the combobox valuesmant
                    matches = [item for item in c2['values'] if search_term.lower() in item.lower()]
                    if matches:
                        c2['values'] = matches
                        c2.current(0)
                    else:
                        c2['values'] = nomenclatura
                        c2.current(0)

                # label
                etiquetaCuenta = tk.Label(ventana, text="Buscar Cuenta", font=("Courier", 14, 'bold'))
                etiquetaCuenta.place(x=50,y=550)
                etiquetaCuenta.configure(bg='#003153', fg='#DADEE7')

                # Trace the variable
                search_var.trace_add('write', update_combobox)

                # Create the search entry
                search_entry = tk.Entry(ventana, textvariable=search_var, font=("Courier", 14),width=55, bg='#DADEE7', fg='#003153')
                search_entry.place(x=210,y=550)
    
                #request focus
                search_entry.focus()
                
                # NOMENCLATURA DE CUENTAS          
                c2 = ttk.Combobox(ventana,values=nomenclatura,font=("Courier", 14),width=80)
                c2.current(0)
                c2.place(x=50,y=520)
                c2['state'] = 'readonly'

                # label
                etiquetaConcepto = tk.Label(ventana, text="Concepto", font=("Courier", 14, 'bold'))
                etiquetaConcepto.place(x=50,y=590)
                etiquetaConcepto.configure(bg='#003153', fg='#DADEE7')

                c1 = ttk.Combobox(ventana,font=("Courier", 16),width=35)
                c1['values']=('Varios','Medicamentos Genéricos','Vehículos','Bebidas','Importaciones','Consumo Combustible','Compra Activos Fijos','Importacion Activos Fijos','Importaciones de C.A.','Compras no Afectas (29-89)','Combustible para Venta')
                c1.current(0)
                c1.place(x=200,y=590)
                c1['state'] = 'readonly'

                ventana.wait_window(ventana)
            
            except Exception as e:
                print(e)
                # messagebox.showwarning("Error","No se pudo extraer informacion del archivo : " + str(e))
                continue
        operaciones_ordenadas = sorted(compras,key=lambda Compra: Compra.fechaObj)
        file_export_csv = path_DataOutputCompras + "\\" + nombreReceptor + "_COMPRAS_" + periodo + ".csv"

        with open(file_export_csv,'w') as f:
            headersCompras = 'nit,razon,peqcont,retenerisr,serie,numero,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descrip,tipodoc,centro,ctagasto,sucursal,isrreten,fchaisr,visrreten,isrtipo,isrconcep,isrbase\n'
            f.write(headersCompras)
            for compra in operaciones_ordenadas:
                strCompra = '{}\n'.format(compra)
                f.write(strCompra)
            f.close()
        messagebox.showinfo("Correcto","Se analizaron las facturas, se encuentraen en carpeta Documentos/FEL-A-SICOFI")

        try:
            #create log file with null files
            file_log = path_DataOutputVentas + "\\" + nombreReceptor +"_"+mesYear + "_log.txt"
            with open(file_log,'w') as f:
                if len(nulls) == 0 and len(download_errors) == 0:
                    f.write('No hay archivos nulos')
                else:
                    f.write('!! ERRORES deberá descargar de nuevo los siguientes archivos o ingresarlos manualmente a SICOFI !! \n\n')
                    for null in nulls:
                        f.write('Error Nulo descargar de nuevo [ ' + null + ' ]\n')
                    for download_error in download_errors:
                        f.write('Error: Descargar de nuevo SAT dte [ ' + download_error + ' ]\n')
                f.close()
            
            # abrir archivo log
            os.startfile(file_log)
        except Exception as e:
            print(e)
            messagebox.showwarning("Error","No se pudo crear archivo log " + str(e))

        compras_ordenadas_no_procesadas = sorted(compras_no_procesadas,key=lambda Compra: Compra.fechaObj)
        file_export_csv_not_processed = path_DataOutputCompras + "\\" + nombreReceptor + "_COMPRAS_NO_PROCESADAS_" + periodo + ".csv"

        with open(file_export_csv_not_processed,'w') as f:
            headersCompras = 'nit,razon,peqcont,retenerisr,serie,numero,fecha,periodo,serv_af,serv_naf,serv_afe,serv_nafe,comp_af,comp_naf,comp_afe,comp_nafe,otros,concepto,descrip,tipodoc,centro,ctagasto,sucursal,isrreten,fchaisr,visrreten,isrtipo,isrconcep,isrbase\n'
            f.write(headersCompras)
            # escribir mensaje en archivo COMPRA NO PROCESADAS HACERLO DE FORMA MANUAL
            f.write('COMPRAS NO PROCESADAS HACERLO DE FORMA MANUAL | COMPRAS NO PROCESADAS HACERLO DE FORMA MANUAL | COMPRAS NO PROCESADAS HACERLO DE FORMA MANUAL\n')
            for compra in compras_ordenadas_no_procesadas:
                strCompra = '{}\n'.format(compra)
                f.write(strCompra)
            f.close()
        webbrowser.open(path_DataOutputCompras)
        # intentar abrir los archivos file_export_csv y file_export_csv_not_processed
        try :
            os.startfile(file_export_csv)
            os.startfile(file_export_csv_not_processed)
        except Exception as e:
            messagebox.showwarning("Error","No se pudo abrir los archivos generados" + str(e))

    except Exception as e:
        print(e)
        messagebox.showwarning("Error","No se abrio ningun archivo " + str(e))

def extraerInfoVentas():
    try:
        path_downloads = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads')     
        file_names = filedialog.askopenfilenames(initialdir=path_downloads,title = "Selecciona uno o varios archivos")
        list_file_names = list(file_names)
        operaciones=[]
        nulls=[]
        download_errors = []
        ns = {'DTE':'http://www.sat.gob.gt/dte/fel/0.2.0'}
        for file_name in list_file_names:
            #si archivo solo contiene la palabra "null" no se procesa
            if os.path.getsize(file_name) == 4:
                #extract file name without extension
                file_name_export = os.path.basename(file_name).split(".")[0]
                nulls.append(file_name_export)
                continue
            try:
                try:
                    tree = ET.parse(file_name,parser=ET.XMLParser(encoding="utf-8"))
                except Exception as e:
                    print(e)
                    messagebox.showwarning(f"Error","El documento {file_name_export} se debe descargar de nuevo del portal de la SAT")
                    download_errors.append(file_name_export)
                    continue

                ET.indent(tree, space="\t", level=0)
                tree.write(file_name, encoding="utf-8")     
                tree = ET.parse(file_name ,parser=ET.XMLParser(encoding="utf-8"))
                root = tree.getroot()  

                try:
                    DATOS_Emision = root.find('DTE:SAT',ns).find('DTE:DTE',ns).find('DTE:DatosEmision',ns)
                    DATOS_Certificacion = root.find('DTE:SAT',ns).find('DTE:DTE',ns).find('DTE:Certificacion',ns)
                except:
                    continue

                nombreEmisor = DATOS_Emision.find('DTE:Emisor',ns).get('NombreEmisor')
                nombreEmisor = re.sub(r'[<>:"/\\|?*]', '_', nombreEmisor)
                nombreEmisor = nombreEmisor.replace(",","")
                nombreEmisor = nombreEmisor.replace(" ","_")
                nombreEmisor = nombreEmisor.replace(".","_")
                nombreEmisor = nombreEmisor.replace("-","_")
                nombreEmisor = nombreEmisor.replace("¿","")

                Datosgen = DATOS_Emision.find('DTE:DatosGenerales',ns)
                fecha = Datosgen.get('FechaHoraEmision')
                fecha = fecha.split("-")
                anio = fecha[0]
                mes = fecha[1]
                dia = fecha[2].split('T')[0]
                fecha = ('{}/{}/{}'.format(dia,mes,anio))
                fecha_obj = datetime.strptime(fecha, '%d/%m/%Y').date()

                #certificacion
                nitReceptor = DATOS_Emision.find('DTE:Receptor',ns).get('IDReceptor')
                NombreReceptor = DATOS_Emision.find('DTE:Receptor',ns).get('NombreReceptor')

                if NombreReceptor.count(',') > 1:
                    NombreReceptor = NombreReceptor.replace(","," ")
                    NombreReceptor = re.sub(r'[<>:"/\\|?*]', '_', NombreReceptor)
                else:
                    NombreReceptor = NombreReceptor.replace(",","")
                    
                noDTE  = DATOS_Certificacion.find('DTE:NumeroAutorizacion',ns).get('Numero')
                serie = DATOS_Certificacion.find('DTE:NumeroAutorizacion',ns).get('Serie')
                
                #items
                items = DATOS_Emision.find('DTE:Items',ns)
                for item in items:
                    tipo = item.get('BienOServicio')
                    if tipo == 'B':
                        tipo = 'Bien'
                    else:
                        tipo = 'Servicio'

                    cant = item.find('DTE:Cantidad',ns).text
                    cant = int(float(cant))
                    descripcion = item.find('DTE:Descripcion',ns).text
                    descripcion = descripcion.replace('\n','')
                    descripcion = descripcion.replace('\t','')
                    descripcion = descripcion.replace(',','')
                    precioUnit = float(item.find('DTE:PrecioUnitario',ns).text)
                    total = float(item.find('DTE:Total',ns).text)       
                    temp = Item(cant,descripcion,precioUnit,total,tipo)
                    operaciones.append(Operacion(fecha_obj,nitReceptor,NombreReceptor,temp,noDTE,serie)) 
            except Exception as e:
                print(e)
                messagebox.showwarning("Error","No se pudo extraer informacion del archivo : " + str(e))
                continue
        operaciones_ordenadas = sorted(operaciones,key=lambda Operacion: Operacion.fecha)
        file_export_csv = path_DataOutputVentas + "\\filetemp.csv"
        mesandyear =''
        
        with open(file_export_csv,'w') as f:
            headers = 'Fecha,Serie,Factura,NIT,Nombre,Unidad,Descripcion,Total\n'
            f.write(headers)
            for operacion in operaciones_ordenadas:
                fechastr = operacion.fecha.strftime("%d/%m/%Y")
                cant = str(operacion.item.cantidad)
                descripcion = operacion.item.descripcion
                precioUnit = str(operacion.item.precio)
                total = str(operacion.item.total)

                NombreReceptor = operacion.nombreReceptor
                nitReceptor = operacion.nitReceptor
                NoDte = operacion.noDTE
                serie = operacion.serie

                strOperacion = '{},{},{},{},{},{},{},{}\n'.format(fechastr,serie,NoDte,nitReceptor,NombreReceptor,cant,descripcion,total)
                f.write(strOperacion) 
                mesandyear = operacion.fecha.strftime("%m-%Y")     
            #cerrar archivo
            f.close()
      
        try:
            #create log file with null files
            file_log = path_DataOutputVentas + "\\" + nombreEmisor +"_"+mesandyear + "_log.txt"
            with open(file_log,'w') as f:
                if len(nulls) == 0 and len(download_errors) == 0:
                    f.write('No hay archivos nulos')
                else:
                    f.write('!! ERRORES deberá descargar de nuevo los siguientes archivos o ingresarlos manualmente a SICOFI !! \n\n')
                    for null in nulls:
                        f.write('Error Nulo descargar de nuevo [ ' + null + ' ]\n')
                    for download_error in download_errors:
                        f.write('Error: Descargar de nuevo SAT dte [ ' + download_error + ' ]\n')
                f.close()
            
            # abrir archivo log
            os.startfile(file_log)
        except Exception as e:
            print(e)
            messagebox.showwarning("Error","No se pudo crear archivo log " + str(e))

        pf = pd.read_csv(file_export_csv,encoding='utf-8')
        #column Unidad type int
        pf['Unidad'] = pf['Unidad'].astype(int)
    
        #column Factura type text 
        pf['Factura'] = pf['Factura'].astype(str)

        #column NIT type text
        pf['NIT'] = pf['NIT'].astype(str)

        #column Total type Quetzal GTQ
        pf['Total'] = pf['Total'].astype(float)
 
        try:
        
            pf.style.set_properties(**{'text-align': 'center'})
        except Exception as e:
            print(e)
        # maneja la excepción aquí


        file_export_xlsx = path_DataOutputVentas + "\\" + nombreEmisor +"_"+mesandyear + "_DETALLE_FACTURA.xlsx"
        writer = pd.ExcelWriter(file_export_xlsx)
        pf.to_excel(writer,sheet_name='REPORTE INVENTARIO POR FACTURA',index=False,header=True,na_rep='NaN')

        try:
            format_center = writer.book.add_format()
            format_center.set_align('center')
            format_center.set_align('vcenter')
            format_center.set_border()

            format_left = writer.book.add_format()
            format_left.set_align('left')
            format_left.set_align('vcenter')
            format_left.set_border()
        except Exception as e:
            print(e)
            
        #columna Fecha
        col_fecha = pf.columns.get_loc('Fecha')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_fecha,col_fecha,12,format_center)

        #columna Serie
        col_serie = pf.columns.get_loc('Serie')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_serie,col_serie,10.5,format_center)


        #columna Factura
        col_factura = pf.columns.get_loc('Factura')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_factura,col_factura,12,format_center)
     

        #columna NIT
        col_nit = pf.columns.get_loc('NIT')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_nit,col_nit,10,format_center)
  
        #columna Nombre
        col_nombre = pf.columns.get_loc('Nombre')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_nombre,col_nombre,60,format_left)
        
        #columna Unidad
        col_unidad = pf.columns.get_loc('Unidad')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_unidad,col_unidad,7,format_center)

        #columna Descripcion
        col_descripcion = pf.columns.get_loc('Descripcion')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_descripcion,col_descripcion,60,format_left)

        #format Quetzal GTQ
        format_gtq = writer.book.add_format({'num_format': 'Q #,##0.00'})
        format_gtq.set_align('right')
        format_gtq.set_align('vcenter')
        format_gtq.set_border()

        #columna Total
        col_total = pf.columns.get_loc('Total')
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].set_column(col_total,col_total,12,format_gtq)
        
        #Freeze pane on the top row.
        writer.sheets['REPORTE INVENTARIO POR FACTURA'].freeze_panes(1, 0)

        
        writer.close()
        
        #delete file temp
        os.remove(file_export_csv)
        messagebox.showinfo("Correcto","Se analizaron los archivos correctamente")
    except Exception as e:
        print(e)
        messagebox.showwarning("Error","No se abrio ningun archivo " + str(e))

def prepareFileExcel (file_name_xls):
    try:
        if file_name_xls.endswith('.xls'):

            file_name_xlsx = file_name_xls.replace('.xls','.xlsx')
            wb_xls = xlrd.open_workbook(file_name_xls)
            wb_xlsx = openpyxl.Workbook()
            ws_xlsx = wb_xlsx.active
            for sheet_name in wb_xls.sheet_names():
                sh_xls = wb_xls.sheet_by_name(sheet_name)
                for row in range(sh_xls.nrows):
                    for col in range(sh_xls.ncols):
                        c = sh_xls.cell(row,col)
                        c.value = str(c.value).replace('Ã¡','á')
                        c.value = str(c.value).replace('Ã©','é')
                        c.value = str(c.value).replace('Ã­','í')
                        c.value = str(c.value).replace('Ã³','ó')
                        c.value = str(c.value).replace('Ãº','ú')

                        ws_xlsx.cell(row=row+1, column=col+1).value = c.value
            
            wb_xlsx.save(file_name_xlsx)

            return file_name_xlsx
        else:
            file_name_xlsx = file_name_xls
            wb_xlsx = openpyxl.load_workbook(file_name_xlsx)
            ws_xlsx = wb_xlsx.active
            for row in range(ws_xlsx.max_row):
                for col in range(ws_xlsx.max_column):
                    c = ws_xlsx.cell(row=row+1, column=col+1)
                    c.value = str(c.value).replace('Ã¡','á')
                    c.value = str(c.value).replace('Ã©','é')
                    c.value = str(c.value).replace('Ã­','í')
                    c.value = str(c.value).replace('Ã³','ó')
                    c.value = str(c.value).replace('Ãº','ú')
            wb_xlsx.save(file_name_xlsx)
            return file_name_xlsx

    except:
        messagebox.showwarning("Error","No se pudo convertir el archivo")

def truncar(numero, decimales):
    factor = 10.0 ** decimales
    return math.floor(numero * factor) / factor


carga_Menu()